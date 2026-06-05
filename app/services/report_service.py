"""
Report generation service for attendance data.
Supports Excel export and statistical summaries.
"""

import logging
import io
from datetime import date, datetime, timedelta
from typing import List, Optional, Dict

from sqlalchemy.orm import Session
from sqlalchemy import func, and_

from app.models.attendance import AttendanceRecord
from app.models.employee import Employee
from app.models.department import Department

logger = logging.getLogger(__name__)


class ReportService:
    """Generate attendance reports and statistics."""

    def get_daily_report(
        self, db: Session, report_date: date
    ) -> List[Dict]:
        """Get attendance report for a specific date."""
        employees = (
            db.query(Employee)
            .filter(Employee.is_active == True)
            .order_by(Employee.employee_code)
            .all()
        )

        records_map = {}
        records = (
            db.query(AttendanceRecord)
            .filter(AttendanceRecord.attendance_date == report_date)
            .all()
        )
        for r in records:
            records_map[r.employee_id] = r

        report = []
        for emp in employees:
            record = records_map.get(emp.id)
            dept_name = emp.department.name if emp.department else ""

            report.append({
                "employee_code": emp.employee_code,
                "employee_name": emp.full_name,
                "department": dept_name,
                "check_in": (
                    record.check_in_time.strftime("%H:%M:%S")
                    if record and record.check_in_time
                    else ""
                ),
                "check_out": (
                    record.check_out_time.strftime("%H:%M:%S")
                    if record and record.check_out_time
                    else ""
                ),
                "status": record.status if record else "ABSENT",
                "source": record.source if record else "",
                "note": record.note if record else "",
            })

        return report

    def get_monthly_summary(
        self,
        db: Session,
        year: int,
        month: int,
    ) -> List[Dict]:
        """Get monthly attendance summary per employee."""
        # Date range for the month
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)

        employees = (
            db.query(Employee)
            .filter(Employee.is_active == True)
            .order_by(Employee.employee_code)
            .all()
        )

        summary = []
        for emp in employees:
            records = (
                db.query(AttendanceRecord)
                .filter(
                    AttendanceRecord.employee_id == emp.id,
                    AttendanceRecord.attendance_date >= start_date,
                    AttendanceRecord.attendance_date < end_date,
                )
                .all()
            )

            present = sum(1 for r in records if r.status == "PRESENT")
            late = sum(1 for r in records if r.status == "LATE")
            half_day = sum(1 for r in records if r.status == "HALF_DAY")

            # Calculate working days in the month (Mon-Fri)
            working_days = 0
            current = start_date
            while current < end_date:
                if current.weekday() < 5:
                    working_days += 1
                current += timedelta(days=1)

            absent = working_days - present - late - half_day

            dept_name = emp.department.name if emp.department else ""

            summary.append({
                "employee_id": emp.id,
                "employee_code": emp.employee_code,
                "employee_name": emp.full_name,
                "department": dept_name,
                "working_days": working_days,
                "present_days": present,
                "late_days": late,
                "half_days": half_day,
                "absent_days": max(0, absent),
                "attendance_rate": round(
                    (present + late + half_day) / max(working_days, 1) * 100, 1
                ),
            })

        return summary

    def export_to_excel(
        self,
        db: Session,
        year: int,
        month: int,
    ) -> bytes:
        """
        Export monthly attendance report to Excel bytes.

        Returns:
            Excel file as bytes (can be sent as StreamingResponse).
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        summary = self.get_monthly_summary(db, year, month)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance {year}-{month:02d}"

        # Title
        ws.merge_cells("A1:J1")
        title_cell = ws["A1"]
        title_cell.value = f"BÁO CÁO CHẤM CÔNG THÁNG {month:02d}/{year}"
        title_cell.font = Font(name="Arial", size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Headers
        headers = [
            "STT", "Mã NV", "Họ tên", "Phòng ban",
            "Ngày công", "Đúng giờ", "Đi trễ", "Nửa ngày",
            "Vắng", "Tỷ lệ (%)"
        ]
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        for idx, row_data in enumerate(summary, 1):
            row = idx + 3
            values = [
                idx,
                row_data["employee_code"],
                row_data["employee_name"],
                row_data["department"],
                row_data["working_days"],
                row_data["present_days"],
                row_data["late_days"],
                row_data["half_days"],
                row_data["absent_days"],
                row_data["attendance_rate"],
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = Font(name="Arial", size=10)
                cell.border = thin_border
                if col >= 5:
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        column_widths = [5, 10, 25, 20, 10, 10, 10, 10, 10, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[
                chr(64 + col) if col <= 26 else "A"
            ].width = width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def get_dashboard_stats(self, db: Session) -> Dict:
        """Get real-time dashboard statistics."""
        today = date.today()

        total_active = (
            db.query(Employee).filter(Employee.is_active == True).count()
        )

        today_records = (
            db.query(AttendanceRecord)
            .filter(AttendanceRecord.attendance_date == today)
            .all()
        )

        checked_in = len(today_records)
        late = sum(1 for r in today_records if r.status == "LATE")
        on_time = sum(1 for r in today_records if r.status == "PRESENT")

        # Recent activities (last 10)
        recent = (
            db.query(AttendanceRecord)
            .join(Employee)
            .filter(AttendanceRecord.attendance_date == today)
            .order_by(AttendanceRecord.check_in_time.desc())
            .limit(10)
            .all()
        )

        recent_activities = []
        for r in recent:
            emp = r.employee
            recent_activities.append({
                "employee_name": emp.full_name if emp else "Unknown",
                "employee_code": emp.employee_code if emp else "",
                "action": "Check-in",
                "time": (
                    r.check_in_time.strftime("%H:%M:%S")
                    if r.check_in_time
                    else ""
                ),
                "status": r.status,
            })

        return {
            "total_employees": total_active,
            "checked_in_today": checked_in,
            "late_today": late,
            "on_time_today": on_time,
            "absent_today": max(0, total_active - checked_in),
            "recent_activities": recent_activities,
        }
